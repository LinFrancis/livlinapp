"""Excel export v7.0 — LivLin · ERP/HRP, escala 0-10."""
import io, json, base64, tempfile, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from utils.petal_content import (
    PETAL_DESC, PETAL_ICONS, IPR_SCALE, IPR_WHAT_IS, IPR_OBS_VS_POT,
    LIVLIN_URL, LIVLIN_TAGLINE, LIVLIN_DESC, LIVLIN_MODULES,
    LIVLIN_SERVICES_PITCH, LIVLIN_CLOSING, REGENERATION_FRAMEWORK, GLOBAL_REFS,
    TAO_REFLEXION_SHORT, TAO_INVITACION
)

C1="1B4332"; C2="2D6A4F"; C3="40916C"; C4="52B788"; C5="D8F3DC"
WHITE="FFFFFF"; GRAY="F5F5F5"; GOLD="FFF8DC"; AMBER="FFF3E0"; AMBER_D="2D6A4F"

def _f(h): return PatternFill("solid", fgColor=h)
def _s(c=C2): return Side(style="thin", color=c)
THIN = Border(left=_s(), right=_s(), top=_s(), bottom=_s())
BOT  = Border(bottom=Side(style="medium", color=C1))

def _cs(ws, r, c, val="", bg=WHITE, fg=C1, bold=False, size=9,
        ha="left", wrap=True, mc=None, border=None, italic=False, url=None):
    if isinstance(val, (list,dict)): val=str(val)
    if val is None: val=""
    cell = ws.cell(row=r, column=c, value=str(val)[:800])
    cell.fill=_f(bg); cell.font=Font(bold=bold,color=fg,size=size,italic=italic,
                                      name="Calibri",underline="single" if url else None)
    cell.alignment=Alignment(horizontal=ha,vertical="top",wrap_text=wrap)
    if mc: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+mc-1)
    if border: cell.border=border
    if url: cell.hyperlink=url
    return cell

def _H(ws,r,txt,bg=C1,fg=WHITE,span=6,h=30,sz=13):
    _cs(ws,r,1,txt,bg=bg,fg=fg,bold=True,size=sz,ha="center",mc=span)
    ws.row_dimensions[r].height=h; return r+1

def _h(ws,r,txt,bg=C2,fg=WHITE,span=6,h=22,sz=10):
    _cs(ws,r,1,txt,bg=bg,fg=WHITE,bold=True,size=sz,mc=span)
    ws.row_dimensions[r].height=h; return r+1

def _expl(ws,r,txt,span=6,h=None):
    _cs(ws,r,1,txt,bg=AMBER,fg=AMBER_D,italic=True,size=8,mc=span,wrap=True)
    ws.row_dimensions[r].height=h or max(18,min(len(txt)//5,55))
    return r+1

def _row(ws,r,lbl,val,h=20):
    _cs(ws,r,1,lbl,bg=C5,bold=True,size=9,mc=2,border=THIN)
    _cs(ws,r,3,str(val) if val else "",bg=WHITE,fg=C2,size=9,mc=4,wrap=True,border=BOT)
    ws.row_dimensions[r].height=h; return r+1

def _sp(ws,r,span=6):
    _cs(ws,r,1,"",bg=C5,mc=span); ws.row_dimensions[r].height=5; return r+1

def _ref(ws,r,auth,title,url,span=6):
    _cs(ws,r,1,auth,bg=GRAY,bold=True,size=8,mc=2,border=THIN)
    _cs(ws,r,3,title,bg=GRAY,size=8,mc=3,wrap=True,border=BOT)
    _cs(ws,r,6,url,bg=GRAY,fg="1565C0",size=8,url=url,border=BOT)
    ws.row_dimensions[r].height=18; return r+1

def _wcol(ws,widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

def _score_lv(n):
    if n==0: return "○ Sin inicio"
    if n<=2: return "🌱 Semilla"
    if n<=5: return "🌿 Brote"
    if n<=9: return "🌳 Crecimiento"
    return "✨ Abundancia"

def _load_petalos():
    jf = Path(__file__).parent.parent/"data"/"petalos_regeneracion_urbana.json"
    try:
        with open(jf,encoding="utf-8") as f: return json.load(f)["petalos"]
    except: return []

def _add_logo(ws, row=1, col=1, size_cm=2.5):
    try:
        from utils.logo_b64 import LOGO_B64
        tmp_path = "/tmp/livlin_logo.png"
        if not os.path.exists(tmp_path):
            logo_bytes = base64.b64decode(LOGO_B64)
            with open(tmp_path, "wb") as _fh2: _fh2.write(logo_bytes)
        img = XLImage(tmp_path)
        px = int(size_cm * 37.8)
        img.width = px; img.height = px
        ws.add_image(img, ws.cell(row=row, column=col).coordinate)
        return True
    except Exception as e:
        print(f"[logo] {e}"); return False


def generate_excel(data: dict) -> bytes:
    wb = Workbook()
    petalos = _load_petalos()
    ipr_obs = data.get("ipr_obs", [])
    ipr_new = data.get("ipr_new", [])
    facilitador = data.get("proyecto_facilitador","")

    # Compute ERP/HRP
    from utils.scoring import (compute_erp, compute_hrp, compute_domain_scores,
        compute_domain_scores_total, compute_cross_module_score, CROSS_MODULE_DETAIL,
        compute_synthesis_potentials, compute_synthesis_potentials_obs,
        score_label, brecha_valor, brecha_texto, _score_to_level, get_interp_text)
    erp = compute_erp(data)
    hrp = compute_hrp(data)
    brecha = brecha_valor(erp, hrp)
    brecha_txt = brecha_texto(erp, hrp)
    label_erp, _ = score_label(erp)
    label_hrp, _ = score_label(hrp)
    domain_obs = compute_domain_scores(data)
    domain_tot = compute_domain_scores_total(data)
    cross = compute_cross_module_score(data)
    pot_erp = compute_synthesis_potentials_obs(data)
    pot_hrp = compute_synthesis_potentials(data)

    # ═══ HOJA 0 — PRESENTACIÓN ═══
    ws0=wb.active; ws0.title="🌿 Presentación"
    _wcol(ws0,[26,16,20,18,20,16])
    _add_logo(ws0, row=1, col=1, size_cm=2.2)
    ws0.merge_cells("B1:F1")
    ws0["B1"]="🌿 LivLin · Indagación Regenerativa v7.0"
    ws0["B1"].fill=_f(C1); ws0["B1"].font=Font(color=WHITE,bold=True,size=18,name="Calibri")
    ws0["B1"].alignment=Alignment(horizontal="center",vertical="center")
    ws0.row_dimensions[1].height=55
    ws0.merge_cells("A2:F2")
    ws0["A2"]=f"Potencial para una vida regenerativa · www.livlin.cl"
    ws0["A2"].fill=_f(C2); ws0["A2"].font=Font(color=WHITE,size=11,italic=True); ws0["A2"].alignment=Alignment(horizontal="center")
    ws0.row_dimensions[2].height=26
    r=3; r=_sp(ws0,r)

    r=_h(ws0,r,"📋 Información del Diagnóstico")
    for lbl,key in [("Espacio","proyecto_nombre"),("Contacto","proyecto_cliente"),
                    ("Ciudad","proyecto_ciudad"),("Fecha","proyecto_fecha"),
                    ("Tipo de espacio","proyecto_tipo_espacio"),("Mascotas","proyecto_mascotas")]:
        v = data.get(key,"")
        if v: r=_row(ws0,r,lbl,v)
    if facilitador: r=_row(ws0,r,"Facilitador/a",facilitador)
    r=_sp(ws0,r)

    # ERP / HRP summary
    r=_h(ws0,r,"📊 Estado Regenerativo Presente (ERP) y Horizonte Regenerativo Potencial (HRP)")
    r=_expl(ws0,r,IPR_WHAT_IS,h=50)
    r=_expl(ws0,r,IPR_OBS_VS_POT,h=45)
    r=_sp(ws0,r)
    r=_row(ws0,r,"🌍 ERP (Estado presente)",f"{erp}/10 — {label_erp}")
    r=_row(ws0,r,"🌱 HRP (Horizonte potencial)",f"{hrp}/10 — {label_hrp}")
    r=_row(ws0,r,"🌀 Brecha (HRP − ERP)",f"{brecha} pts — {brecha_txt}")
    r=_sp(ws0,r)

    r=_h(ws0,r,"🔢 Escala de niveles (0-10)")
    for lvl,n,_,meaning in IPR_SCALE:
        bg=C5 if r%2==0 else WHITE
        _cs(ws0,r,1,lvl,bg=bg,bold=True,size=9,border=THIN)
        _cs(ws0,r,2,n,bg=bg,size=9,ha="center",border=THIN)
        _cs(ws0,r,3,meaning,bg=bg,size=9,mc=4,wrap=True,border=BOT)
        ws0.row_dimensions[r].height=28; r+=1
    r=_sp(ws0,r)

    r=_h(ws0,r,"📚 Referencias")
    for auth,title,url in GLOBAL_REFS:
        r=_ref(ws0,r,auth,title,url)

    # ═══ HOJA 1 — RESUMEN ERP/HRP ═══
    ws1=wb.create_sheet("📊 Resumen ERP-HRP")
    _wcol(ws1,[26,14,14,12,14,14])
    _add_logo(ws1,1,1,1.8)
    ws1.merge_cells("B1:F1")
    ws1["B1"]="📊 RESUMEN — ERP y HRP"
    ws1["B1"].fill=_f(C1); ws1["B1"].font=Font(color=WHITE,bold=True,size=13)
    ws1["B1"].alignment=Alignment(horizontal="center",vertical="center")
    ws1.row_dimensions[1].height=45
    r=2
    r=_expl(ws1,r,"ERP = 80% MFP observado + 20% sub-indicadores M2-6. HRP = 100% MFP proyectado.",h=30)
    r=_sp(ws1,r)
    r=_row(ws1,r,"Espacio",data.get("proyecto_nombre",""))
    r=_row(ws1,r,"🌍 ERP",f"{erp}/10 — {label_erp}")
    r=_row(ws1,r,"🌱 HRP",f"{hrp}/10 — {label_hrp}")
    r=_row(ws1,r,"🌀 Brecha",f"{brecha} pts — {brecha_txt}")
    r=_sp(ws1,r)

    # Pétalos ERP vs HRP
    r=_h(ws1,r,"🌸 Pétalos — ERP vs HRP")
    _cs(ws1,r,1,"Pétalo",bg=C4,bold=True,size=9,mc=2,border=THIN)
    _cs(ws1,r,3,"ERP",bg=C5,bold=True,size=9,ha="center",border=THIN)
    _cs(ws1,r,4,"HRP",bg=AMBER,bold=True,size=9,ha="center",border=THIN)
    _cs(ws1,r,5,"Brecha",bg=C4,bold=True,size=9,ha="center",border=THIN)
    _cs(ws1,r,6,"Nivel ERP",bg=C4,bold=True,size=9,ha="center",border=THIN)
    ws1.row_dimensions[r].height=20; r+=1
    from utils.scoring import PETAL_ORDER
    for i,p in enumerate(PETAL_ORDER):
        icon=PETAL_ICONS[i] if i<len(PETAL_ICONS) else "🌱"
        e=domain_obs[p]; h_=domain_tot[p]; gap=round(h_-e,1)
        bg=C5 if i%2==0 else WHITE
        _cs(ws1,r,1,f"{icon} {p}",bg=bg,bold=True,size=9,mc=2,border=THIN)
        _cs(ws1,r,3,f"{e:.0f}",bg=C5,size=10,ha="center",border=THIN)
        _cs(ws1,r,4,f"{h_:.0f}",bg=AMBER,fg=AMBER_D,size=10,ha="center",border=THIN)
        _cs(ws1,r,5,f"+{gap:.0f}" if gap>0 else "0",bg=bg,size=9,ha="center",border=THIN)
        lv,_=_score_to_level(e)
        _cs(ws1,r,6,lv,bg=bg,size=8,ha="center",border=THIN)
        ws1.row_dimensions[r].height=22; r+=1
    r=_sp(ws1,r)

    # Sub-indicadores
    if cross:
        r=_h(ws1,r,"📈 Sub-indicadores M2-6 (aportan 20% al ERP)")
        for name, info in cross.items():
            r=_row(ws1,r,name,f"{info['score']}/10 — {info['fuente']}")
        r=_sp(ws1,r)
        # Transparency
        r=_h(ws1,r,"🔍 Transparencia: variables y escalas")
        from utils.scoring import CROSS_MODULE_DETAIL
        for name, detail in CROSS_MODULE_DETAIL.items():
            actual = cross.get(name,{}).get("score","—")
            r=_expl(ws1,r,f"{detail.get('icono','')} {name}: {actual}/10 · {detail['formula']}",h=20)
            for vn,vs in detail["variables"]:
                _cs(ws1,r,1,f"  • {vn}: {vs}",bg=WHITE,size=8,mc=6)
                ws1.row_dimensions[r].height=16; r+=1

    # 10 dimensions
    r=_sp(ws1,r)
    r=_h(ws1,r,"🌿 10 Dimensiones — ERP vs HRP")
    for dim in pot_erp:
        e=pot_erp[dim]; h_=pot_hrp.get(dim,0)
        interp_e = get_interp_text(dim, e, "erp")
        interp_h = get_interp_text(dim, h_, "hrp")
        _cs(ws1,r,1,dim,bg=C4,bold=True,size=9,mc=2,border=THIN)
        _cs(ws1,r,3,f"ERP: {e}/10",bg=C5,size=9,ha="center",border=THIN)
        _cs(ws1,r,4,f"HRP: {h_}/10",bg=AMBER,size=9,ha="center",border=THIN)
        _cs(ws1,r,5,f"+{round(h_-e,1)}",bg=WHITE,size=9,ha="center",border=THIN)
        ws1.row_dimensions[r].height=20; r+=1
        if interp_e:
            r=_expl(ws1,r,f"ERP: {interp_e}",h=25)
        if interp_h:
            r=_expl(ws1,r,f"HRP: {interp_h}",h=25)
    r=_sp(ws1,r)
    r=_h(ws1,r,"🌿 LivLin · www.livlin.cl",bg=C2)
    r=_expl(ws1,r,LIVLIN_SERVICES_PITCH,h=55)

    # ═══ HOJA 2 — M1 ═══
    ws2=wb.create_sheet("📋 M1 Intención + Tao")
    _wcol(ws2,[24,14,24,14,12,10])
    r=_H(ws2,1,"📋 MÓDULO 1 — Intención y Contexto + Tao")
    r=_expl(ws2,r,"Horizonte de sentido del diagnóstico.",h=30)
    r=_sp(ws2,r)
    for lbl,key in [("Nombre","proyecto_nombre"),("Contacto","proyecto_cliente"),
                    ("Ciudad","proyecto_ciudad"),("Dirección","proyecto_direccion"),
                    ("Tipo","proyecto_tipo_espacio"),("Superficie (m²)","proyecto_superficie"),
                    ("Habitantes","proyecto_habitantes"),("Mascotas","proyecto_mascotas")]:
        v=data.get(key,"")
        if v: r=_row(ws2,r,lbl,str(v))
    r=_sp(ws2,r)
    r=_h(ws2,r,"🌀 Tao de la Regeneración")
    for lbl,key in [("Sensación","tao_sensacion"),("Deseado","tao_deseado"),
                    ("No deseado","tao_no_deseado"),("Lo que llama","tao_llama"),
                    ("Naturaleza","tao_naturaleza_rel"),("Palabra esencial","tao_palabra_esencial")]:
        v=data.get(key,"")
        if v: r=_row(ws2,r,lbl,str(v))

    # ═══ HOJA 3 — M2-3 ═══
    ws3=wb.create_sheet("🔬 M2-3 Ecología")
    _wcol(ws3,[24,14,24,14,12,10])
    r=_H(ws3,1,"🔬 M2-3 — Observación Ecológica")
    r=_sp(ws3,r)
    for lbl,key in [("Suelo tipo","suelo_tipo"),("Compactación","suelo_compactacion"),
                    ("Materia orgánica","suelo_materia_organica"),("Drenaje","suelo_drenaje"),
                    ("Horas sol","sol_horas"),("Orientación","sol_orientacion"),
                    ("Cultivo m²","cultivo_m2"),("Produce hoy","cultivo_produce_hoy")]:
        v=data.get(key,"")
        if v: r=_row(ws3,r,lbl,str(v))

    # ═══ HOJA 4 — M4-6 ═══
    ws4=wb.create_sheet("🏙️ M4-6 Sistemas")
    _wcol(ws4,[24,14,24,14,12,10])
    r=_H(ws4,1,"🏙️ M4-6 — Contexto, Agua, Energía")
    r=_sp(ws4,r)
    for lbl,key in [("Cuenca","ctx_cuenca"),("Vecinos","ctx_vecinos"),
                    ("Consumo agua","agua_consumo"),("Captación","agua_captacion_lluvia"),
                    ("Grises","agua_grises"),("Fuente energía","ene_fuente"),
                    ("LED","ene_led"),("kWh/día","ene_kwh_dia_calc")]:
        v=data.get(key,"")
        if v: r=_row(ws4,r,lbl,str(v))

    # ═══ HOJA 5 — M7 FLOR ═══
    ws5=wb.create_sheet("🌸 M7 Flor Permacultura")
    _wcol(ws5,[28,18,18,6,28])
    _add_logo(ws5,1,1,1.8)
    ws5.merge_cells("B1:E1")
    ws5["B1"]="🌸 FLOR DE LA PERMACULTURA — ERP / HRP"
    ws5["B1"].fill=_f(C1); ws5["B1"].font=Font(color=WHITE,bold=True,size=13)
    ws5["B1"].alignment=Alignment(horizontal="center",vertical="center")
    ws5.row_dimensions[1].height=45
    r=2
    r=_expl(ws5,r,f"ERP: {erp}/10 — {label_erp} | HRP: {hrp}/10 — {label_hrp} | Brecha: {brecha} pts",span=5,h=25)
    r=_sp(ws5,r,span=5)

    for i,p in enumerate(petalos):
        icon=PETAL_ICONS[i] if i<len(PETAL_ICONS) else "🌱"
        n_o=ipr_obs[i] if i<len(ipr_obs) else 0
        n_n=ipr_new[i] if i<len(ipr_new) else 0
        e_score = domain_obs.get(p["nombre"],0)
        h_score = domain_tot.get(p["nombre"],0)
        lv_e,_=_score_to_level(e_score)
        lv_h,_=_score_to_level(h_score)

        r=_h(ws5,r,f"{icon} {p['nombre']}  ·  ERP {e_score:.0f} → HRP {h_score:.0f}",span=5)
        _cs(ws5,r,1,f"ERP: {n_o} prácticas — {lv_e}",bg=C5,bold=True,size=9,mc=2,border=THIN)
        _cs(ws5,r,3,f"HRP: +{n_n} potencial — {lv_h}",bg=AMBER,fg=AMBER_D,bold=True,size=9,mc=3,border=THIN)
        ws5.row_dimensions[r].height=22; r+=1

        obs_data=data.get(f"petalo_{i}_obs",{})
        new_data=data.get(f"petalo_{i}_pot_new",{})
        otros_obs=data.get(f"petalo_{i}_otros_obs",[])
        otros_new=data.get(f"petalo_{i}_otros_new",[])
        for cat_key in p.get("categorias",{}):
            obs_s=obs_data.get(cat_key,[]); new_s=new_data.get(cat_key,[])
            if not obs_s and not new_s: continue
            _cs(ws5,r,1,cat_key.replace("_"," ").title(),bg=WHITE,bold=True,size=8,border=THIN)
            _cs(ws5,r,2," · ".join(obs_s),bg=C5,size=8,wrap=True,border=THIN)
            _cs(ws5,r,3," · ".join(new_s),bg=AMBER,fg=AMBER_D,size=8,wrap=True,border=THIN)
            ws5.row_dimensions[r].height=max(18,min((len(obs_s)+len(new_s))*12,55)); r+=1
        if otros_obs or otros_new:
            _cs(ws5,r,1,"Otras",bg=WHITE,bold=True,size=8,border=THIN)
            _cs(ws5,r,2," · ".join(otros_obs),bg=C5,size=8,wrap=True,border=THIN)
            _cs(ws5,r,3," · ".join(otros_new),bg=AMBER,fg=AMBER_D,size=8,wrap=True,border=THIN)
            ws5.row_dimensions[r].height=20; r+=1
        r=_sp(ws5,r,span=5)

    # ═══ HOJA 6 — M9 Síntesis + Plan ═══
    ws6=wb.create_sheet("🗺️ M9 Síntesis + Plan")
    _wcol(ws6,[26,14,22,14,14,10])
    r=_H(ws6,1,"🗺️ MÓDULO 9 — Síntesis y Plan de Acción")
    r=_expl(ws6,r,"Diagnóstico → brechas → oportunidades → visión futura.",h=25)
    r=_sp(ws6,r)

    for key,lbl in [("sint_fortalezas","💚 Fortalezas"),("sint_oportunidades","🌱 Oportunidades"),
                    ("sint_limitaciones","⚡ Desafíos"),("sint_quick_wins","🎯 Primeros pasos")]:
        v=data.get(key,"")
        if v: r=_row(ws6,r,lbl,str(v),h=28)
    r=_sp(ws6,r)

    for pk,fase,desc in [
        ("plan_inmediatas","⚡ Fase 1 (0–3 meses)","Acciones inmediatas de bajo costo y alta visibilidad."),
        ("plan_estacionales","🌱 Fase 2 (3–12 meses)","Intervenciones estacionales con planificación."),
        ("plan_estructurales","🌳 Fase 3 (1–5 años)","Transformaciones estructurales de fondo."),
    ]:
        r=_h(ws6,r,fase); r=_expl(ws6,r,desc,h=22)
        v=data.get(pk,"")
        if v:
            if isinstance(v,list):
                for item in v:
                    txt=item.get("titulo","") if isinstance(item,dict) else str(item)
                    if txt: r=_row(ws6,r,"→",txt)
            else:
                _cs(ws6,r,1,str(v),bg=GOLD,fg="5C3D00",size=9,mc=6,wrap=True)
                ws6.row_dimensions[r].height=40; r+=1
        r=_sp(ws6,r)

    r=_h(ws6,r,"🌿 LivLin · www.livlin.cl",bg=C2)
    r=_expl(ws6,r,LIVLIN_CLOSING,h=50)

    # ═══ HOJA 7 — DATOS BIOCLIMÁTICOS ═══
    ws7=wb.create_sheet("🌡️ Datos Bioclimáticos")
    _wcol(ws7,[26,14,24,12,14,10])
    r=_H(ws7,1,"🌡️ Datos Bioclimáticos del Sitio")
    r=_expl(ws7,r,"Datos climáticos generados con APIs: Nominatim/OpenStreetMap + Open-Meteo.",h=35)
    r=_sp(ws7,r)
    for lbl,key in [("Latitud","geo_lat"),("Longitud","geo_lon"),("Ciudad","proyecto_ciudad"),
                    ("Precipitación anual (mm)","agua_prec_anual"),("Mes cálido","clima_mes_caluroso"),
                    ("Mes frío","clima_mes_frio"),("T° máx","clima_t_max_abs"),("T° mín","clima_t_min_abs")]:
        v=data.get(key,"")
        if v: r=_row(ws7,r,lbl,str(v))

    for ws in [ws0,ws1,ws2,ws3,ws4,ws5,ws6,ws7]:
        ws.freeze_panes="A2"; ws.sheet_view.showGridLines=True

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()
